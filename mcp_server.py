# mcp_server.py - THE ABSOLUTELY FINAL CORRECTED VERSION FOR DEPLOYMENT

import os
import json
import asyncio
import logging
from typing import Callable, ClassVar # Added ClassVar as you might have used it in schemas
from fastmcp import FastMCP
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook
from datetime import datetime
from dotenv import load_dotenv
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from twilio.rest import Client
from openpyxl.utils import get_column_letter
from schemas import AppointmentRequest, SmsNotifyRequest, HealthRecordRequest

# ----------- Environment Setup (Simplified for Docker/Cloud) -----------
load_dotenv()

# In a Docker container, these paths will be relative to the /app WORKDIR.
# They will point to files inside the container, which is fine for ephemeral use.
# For persistent storage, you'd integrate with cloud storage (e.g., S3, Google Cloud Storage).
EXCEL_PATH = "appointments.xlsx" # Simplified for container
PDF_PATH = "health_record.pdf"   # Simplified for container

TWILIO_FROM_NUMBER = os.getenv("TWILIO_FROM_NUMBER")
twilio_sid = os.getenv("TWILIO_ACCOUNT_SID")
twilio_token = os.getenv("TWILIO_AUTH_TOKEN")

# Initialize Twilio client only if credentials are provided
client = None
if twilio_sid and twilio_token:
    try:
        client = Client(twilio_sid, twilio_token)
        logger.info("Twilio client initialized in MCP server.")
    except Exception as e:
        logger.error(f"Error initializing Twilio client: {e}. SMS functionality may not work.", exc_info=True)
else:
    logger.warning("Twilio credentials (TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN) not found. SMS functionality will be disabled.")


# ----------- Logging -----------
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ----------- Decorator to register tools -----------
tool_registry = []

def tool(func: Callable):
    tool_registry.append(func)
    return func

# ----------- Utility: Format number -----------
def format_mobile_number_e164(mobile_number, default_country_code="+91"):
    phone = mobile_number.strip().replace(" ", "").replace("-", "")
    if phone.startswith("+"):
        return phone
    if phone.startswith("0"):
        phone = phone[1:]
    if len(phone) == 10 and phone.isdigit():
        return f"{default_country_code}{phone}"
    # This part can be tricky for international numbers. 
    # For a robust solution, consider a dedicated phone number parsing library.
    if len(phone) == 12 and phone.isdigit() and phone.startswith("91"): # Handles 91XXXXXXXXXX if not already +91
        return f"+{phone}"
    if len(phone) == 11 and phone.startswith("91") and phone.isdigit(): # Handles 91XXXXXXXXXX if not already +91
        return f"+{phone}"
    
    # If it's a 10-digit number that didn't start with '+', assume default country code
    if phone.isdigit() and len(phone) == 10 and not phone.startswith('+'):
        return f"{default_country_code}{phone}"

    # If it's just numbers and long enough, try adding default country code if no other pattern matched
    if phone.isdigit() and len(phone) > 7 and not phone.startswith('+'): 
        return f"{default_country_code}{phone}"

    raise ValueError(f"Invalid mobile number for E.164: {mobile_number}")

# ----------- Tool Functions with @tool decorator -----------

@tool
def queue_appointment(input: AppointmentRequest) -> dict:
    """
    Queue a patient for an appointment and log details in an Excel sheet.
    Parameters:
        input (AppointmentRequest): Details of the appointment.
    """
    try:
        logger.info(f"Saving appointment to Excel for {input.patient_name} at {input.appointment_time}...")
        
        # Ensure directory for EXCEL_PATH exists if not just current dir
        excel_dir = os.path.dirname(EXCEL_PATH)
        if excel_dir and not os.path.exists(excel_dir):
            os.makedirs(excel_dir)

        if not os.path.exists(EXCEL_PATH):
            logger.info("Excel file not found. Creating new one.")
            wb = Workbook()
            ws = wb.active
            ws.title = "Appointments"
            ws.append([
                "Patient ID", "Name", "Age", "Gender", "Phone Number",
                "Issue", "Appointment Time", "Queued At"
            ])
            wb.save(EXCEL_PATH)

        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        ws.append([
            input.patient_id, input.patient_name, input.age, input.gender,
            input.phone_number, input.issue, input.appointment_time,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ])
        column_widths = [14, 20, 6, 10, 16, 40, 20, 22]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        wb.save(EXCEL_PATH)
        logger.info(f"✅ Appointment saved at: {EXCEL_PATH}")
        return {
            "status": "success",
            "message": f"Patient {input.patient_name} queued successfully.",
            "path": os.path.abspath(EXCEL_PATH) # Path is within container
        }
    except Exception as e:
        logger.error(f"❌ Failed to save appointment: {e}", exc_info=True)
        return {"status": "error", "message": str(e)}

@tool
def send_appointment_sms(input: SmsNotifyRequest) -> dict:
    """
    Send an SMS notification for a scheduled appointment.
    Parameters:
        input (SmsNotifyRequest): Details for the SMS notification.
    """
    if not client:
        return {"status": "error", "message": "Twilio client not initialized. SMS disabled."}
    if not TWILIO_FROM_NUMBER:
        return {"status": "error", "message": "TWILIO_FROM_NUMBER environment variable is not set. SMS disabled."}
    
    try:
        logger.info(f"Sending SMS notification to {input.phone_number} for {input.patient_name}...")
        to_number = format_mobile_number_e164(input.phone_number)
        body = f"Hello {input.patient_name}, your appointment is scheduled at {input.appointment_time}."
        message = client.messages.create(
            body=body,
            from_=TWILIO_FROM_NUMBER,
            to=to_number
        )
        logger.info(f"✅ SMS sent to {to_number} (SID: {message.sid})")
        return {
            "status": "sent",
            "message": f"SMS sent to {to_number}",
            "sms": body,
            "sid": message.sid
        }
    except Exception as e:
        logger.error(f"❌ SMS sending failed: {e}", exc_info=True)
        return {"status": "error", "message": str(e)}

@tool
def generate_health_record(input: HealthRecordRequest) -> dict:
    """
    Generate a PDF health record for the patient.
    Parameters:
        input (HealthRecordRequest): Patient's health record details.
    """
    try:
        logger.info(f"Generating health record PDF for {input.patient_name}...")

        pdf_dir = os.path.dirname(PDF_PATH)
        if pdf_dir and not os.path.exists(pdf_dir):
            os.makedirs(pdf_dir)

        c = canvas.Canvas(PDF_PATH, pagesize=letter)
        c.setFont("Helvetica", 12)
        y = 750
        line_gap = 20
        entries = [
            f"Patient Name: {input.patient_name}",
            f"Age: {input.age}",
            f"Gender: {input.gender}",
            f"Phone Number: {input.phone_number}",
            f"Symptoms: {input.symptoms}",
            f"Duration: {input.duration}",
            f"Chronic Conditions: {input.chronic_conditions}",
            f"Family History: {input.family_history}",
            f"Diagnosis: {input.diagnosis}",
            f"Prescriptions: {input.prescriptions}"
        ]
        for line in entries:
            c.drawString(30, y, line)
            y -= line_gap
            if y < 50: # Add page break logic for long reports
                c.showPage()
                c.setFont("Helvetica", 12)
                y = 750

        c.save()
        logger.info(f"✅ PDF generated at: {PDF_PATH}")
        return {
            "status": "success",
            "message": "Health record PDF generated successfully.",
            "path": os.path.abspath(PDF_PATH) # Path is within container
        }
    except Exception as e:
        logger.error(f"❌ Failed to generate PDF: {e}", exc_info=True)
        return {"status": "error", "message": str(e)}

# ----------- MCP Init & Run -----------

mcp = FastMCP(
    name="AI-Health-Multilingual-Voice-agent",
    tools=tool_registry # tool_registry now holds the functions from @tool decorator
)

# CRITICAL ADDITION: Expose the FastMCP app instance directly for Uvicorn
app = mcp.app # <--- THIS IS THE KEY LINE FOR DEPLOYMENT!

# This block is for local debugging/running directly, not for Dockerized deployment
if __name__ == "__main__":
    logger.info("🚀 Initializing MCP for AI-Health-Multilingual-Voice-agent...")
    logger.info(f"📂 Excel will be saved to: {EXCEL_PATH}")
    logger.info(f"📂 PDF will be saved to: {PDF_PATH}")
    
    import uvicorn
    logger.info("Running MCP app with Uvicorn (for local testing)...")
    # Use the 'app' variable directly here for local Uvicorn run
    uvicorn.run(app, host="127.0.0.1", port=8000)